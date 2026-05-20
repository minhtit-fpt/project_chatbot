#!/usr/bin/env python3
"""eval/run_excel_test.py — Chạy DeepEval trực tiếp từ file Excel, ghi kết quả vào sheet."""

from __future__ import annotations

import asyncio
import json
import logging
import sys
import io
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import PatternFill, Font

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
logger = logging.getLogger(__name__)

EXCEL_PATH = Path(r"C:\Users\ham48\Downloads\test_chatbot_dienmaythienthu.xlsx")
SHEET_NAME = "Bộ câu hỏi test"

# Cột input (0-indexed)
COL_DANH_MUC = 1
COL_NHAN = 2
COL_LOAI = 3
COL_QUESTION = 4
COL_EXPECTED = 5

# Cột output (H=8 trong openpyxl là column index 8, 1-indexed)
# openpyxl dùng 1-indexed cho columns
OUT_COL_BOT_ANSWER = 8       # H
OUT_COL_SOURCES = 9          # I
OUT_COL_LATENCY = 10         # J
OUT_COL_RELEVANCY = 11       # K
OUT_COL_FAITHFULNESS = 12    # L
OUT_COL_PASS = 13            # M
OUT_COL_RUN_AT = 14          # N
OUT_COL_ERROR = 15           # O

OUTPUT_HEADERS = [
    "bot_answer", "sources", "latency_ms",
    "answer_relevancy", "faithfulness", "pass", "run_at", "error",
]

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")


def make_judge():
    from deepeval.models import DeepEvalBaseLLM

    class GeminiJudge(DeepEvalBaseLLM):
        def __init__(self) -> None:
            import config as cfg
            from google import genai
            self._model_name = "gemini-2.5-flash"
            self._client = genai.Client(api_key=cfg.GEMINI_API_KEY)
            super().__init__()

        def load_model(self):
            return self._client

        def generate(self, prompt: str, schema=None):
            from google.genai import types
            cfg_kwargs: dict = {"temperature": 0}
            if schema is not None:
                cfg_kwargs["response_mime_type"] = "application/json"
                cfg_kwargs["response_schema"] = schema
            response = self._client.models.generate_content(
                model=self._model_name,
                contents=prompt,
                config=types.GenerateContentConfig(**cfg_kwargs),
            )
            text = response.text.strip()
            if schema is not None:
                import json as _j
                try:
                    return schema(**_j.loads(text))
                except Exception:
                    return text
            return text

        async def a_generate(self, prompt: str, schema=None):
            return await asyncio.to_thread(self.generate, prompt, schema)

        def get_model_name(self) -> str:
            return self._model_name

    return GeminiJudge()


def run_metrics(question: str, bot_answer: str, expected: str,
                retrieval_context: list[str], judge) -> dict:
    from deepeval.metrics import AnswerRelevancyMetric, FaithfulnessMetric
    from deepeval.test_case import LLMTestCase

    test_case = LLMTestCase(
        input=question,
        actual_output=bot_answer,
        expected_output=expected or None,
        retrieval_context=retrieval_context or None,
    )

    scores: dict = {"answer_relevancy": "", "faithfulness": ""}
    relevancy_error = False

    # AnswerRelevancy
    try:
        m = AnswerRelevancyMetric(model=judge, threshold=0.5, verbose_mode=False)
        m.measure(test_case)
        scores["answer_relevancy"] = round(m.score, 4)
        relevancy_passed = m.is_successful()
    except Exception as exc:
        logger.warning("AnswerRelevancyMetric lỗi: %s", exc)
        relevancy_passed = False
        relevancy_error = True

    # Faithfulness (tham khảo)
    if retrieval_context:
        try:
            m2 = FaithfulnessMetric(model=judge, threshold=0.5, verbose_mode=False)
            m2.measure(test_case)
            scores["faithfulness"] = round(m2.score, 4)
        except Exception as exc:
            logger.warning("FaithfulnessMetric lỗi: %s", exc)

    scores["pass"] = "ERROR" if relevancy_error else ("TRUE" if relevancy_passed else "FALSE")
    return scores


async def run_all():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Ghi output headers nếu chưa có
    if ws.cell(1, OUT_COL_BOT_ANSWER).value != "bot_answer":
        for i, h in enumerate(OUTPUT_HEADERS, start=OUT_COL_BOT_ANSWER):
            ws.cell(1, i).value = h
            ws.cell(1, i).font = Font(bold=True)

    # Load retriever
    from rag.chat_engine import init_retriever, _get_retriever, answer as engine_answer
    logger.info("Loading index...")
    await init_retriever()

    judge = make_judge()
    logger.info("Judge sẵn sàng: %s", judge.get_model_name())

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        q = row[COL_QUESTION]
        if q:
            rows_data.append(row)

    logger.info("Bắt đầu eval %d câu hỏi\n", len(rows_data))

    results = []
    passed = failed = errors = 0

    for idx, row in enumerate(rows_data, start=1):
        sheet_row = idx + 1  # +1 vì header ở row 1
        question = row[COL_QUESTION]
        expected = row[COL_EXPECTED] or ""
        danh_muc = row[COL_DANH_MUC]
        nhan = row[COL_NHAN]

        logger.info("[%d/%d] [%s/%s] %s", idx, len(rows_data), danh_muc, nhan, question)

        result = {"run_at": datetime.now().isoformat(timespec="seconds"), "error": ""}

        try:
            # Lấy context_docs cho faithfulness
            retriever = await _get_retriever()
            context_docs = await asyncio.to_thread(retriever.search, question)
            retrieval_context = [
                f"{d['title']}: {d['content'][:600]}" for d in context_docs
            ]

            chat_result = await engine_answer(question)
            result["bot_answer"] = chat_result["answer"]
            result["sources"] = ", ".join(d["title"] for d in context_docs)
            result["latency_ms"] = chat_result["latency_ms"]

            scores = run_metrics(question, chat_result["answer"], expected,
                                 retrieval_context, judge)
            result.update(scores)

            status = scores["pass"]
            if status == "TRUE":
                passed += 1
                logger.info("  ✓ pass | relevancy=%s faithfulness=%s latency=%dms",
                            scores["answer_relevancy"], scores.get("faithfulness", "n/a"),
                            chat_result["latency_ms"])
            elif status == "ERROR":
                errors += 1
                logger.info("  ? error metric")
            else:
                failed += 1
                logger.info("  ✗ fail | relevancy=%s faithfulness=%s",
                            scores["answer_relevancy"], scores.get("faithfulness", "n/a"))

        except Exception as exc:
            errors += 1
            result.update({
                "bot_answer": "", "sources": "", "latency_ms": "",
                "answer_relevancy": "", "faithfulness": "",
                "pass": "ERROR", "error": str(exc),
            })
            logger.error("  EXCEPTION: %s", exc)

        results.append((sheet_row, result))

        # Ghi từng row ngay (nếu bị ngắt giữa chừng vẫn có kết quả)
        r = result
        vals = [
            r.get("bot_answer", ""), r.get("sources", ""), r.get("latency_ms", ""),
            r.get("answer_relevancy", ""), r.get("faithfulness", ""),
            r.get("pass", ""), r.get("run_at", ""), r.get("error", ""),
        ]
        for ci, v in enumerate(vals, start=OUT_COL_BOT_ANSWER):
            ws.cell(sheet_row, ci).value = v

        # Màu pass/fail cho cột M
        pass_cell = ws.cell(sheet_row, OUT_COL_PASS)
        if r.get("pass") == "TRUE":
            pass_cell.fill = GREEN_FILL
        elif r.get("pass") == "FALSE":
            pass_cell.fill = RED_FILL
        elif r.get("pass") == "ERROR":
            pass_cell.fill = YELLOW_FILL

    wb.save(EXCEL_PATH)
    logger.info("\nĐã lưu kết quả vào: %s", EXCEL_PATH)

    # ── Báo cáo tổng hợp ─────────────────────────────────────────────────────
    total = len(results)
    logger.info("\n" + "=" * 60)
    logger.info("KẾT QUẢ TỔNG HỢP")
    logger.info("=" * 60)
    logger.info("Tổng câu hỏi : %d", total)
    logger.info("Pass         : %d (%.1f%%)", passed, passed/total*100 if total else 0)
    logger.info("Fail         : %d (%.1f%%)", failed, failed/total*100 if total else 0)
    logger.info("Error metric : %d", errors)
    logger.info("=" * 60)

    # Chi tiết fail
    fail_rows = [(sr, r) for sr, r in results if r.get("pass") == "FALSE"]
    if fail_rows:
        logger.info("\nCÂU HỎI FAIL:")
        for sr, r in fail_rows:
            row_data = rows_data[sr - 2]
            logger.info("  Row %d [%s] %s", sr, row_data[COL_NHAN], row_data[COL_QUESTION])
            logger.info("    relevancy=%s | faithfulness=%s",
                        r.get("answer_relevancy"), r.get("faithfulness"))
            logger.info("    bot: %s", str(r.get("bot_answer", ""))[:120])

    error_rows = [(sr, r) for sr, r in results if r.get("pass") == "ERROR"]
    if error_rows:
        logger.info("\nCÂU HỎI LỖI METRIC:")
        for sr, r in error_rows:
            row_data = rows_data[sr - 2]
            logger.info("  Row %d %s | error: %s",
                        sr, row_data[COL_QUESTION], r.get("error", ""))


if __name__ == "__main__":
    asyncio.run(run_all())
